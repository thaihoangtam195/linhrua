"""Tạo file Excel template mẫu cho chatbot"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()

# ============ SHEET 1: Hướng dẫn ============
ws_guide = wb.active
ws_guide.title = "Hướng dẫn"

# Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

guide_data = [
    ["🤖 HƯỚNG DẪN SỬ DỤNG FILE EXCEL CHO CHATBOT"],
    [""],
    ["📋 CẤU TRÚC FILE:"],
    ["Cột", "Mô tả", "Bắt buộc"],
    ["câu hỏi", "Câu hỏi mẫu mà khách hàng có thể hỏi", "✅ Có"],
    ["câu trả lời", "Câu trả lời tương ứng", "✅ Có"],
    ["hình ảnh", "URL hình ảnh sản phẩm (phải là link public)", "❌ Không"],
    ["từ khóa", "Các từ khóa liên quan (cách nhau bằng dấu phẩy)", "❌ Không"],
    ["danh mục", "Phân loại câu hỏi (Giá cả, Vận chuyển, Sản phẩm...)", "❌ Không"],
    [""],
    ["💡 MẸO VIẾT CÂU HỎI-TRẢ LỜI HIỆU QUẢ:"],
    ["1. Viết nhiều biến thể của cùng 1 câu hỏi (VD: 'giá bao nhiêu', 'bn tiền', 'giá sp')"],
    ["2. Thêm từ khóa để bot nhận diện tốt hơn"],
    ["3. Câu trả lời nên tự nhiên, thân thiện, có emoji"],
    ["4. Luôn kết thúc bằng 'ạ' hoặc 'nha' để thân thiện"],
    [""],
    ["🔤 BOT ĐÃ HIỂU CÁC TỪ VIẾT TẮT:"],
    ["sp=sản phẩm, đh=đơn hàng, vc/ship=vận chuyển, bn=bao nhiêu"],
    ["k/ko=không, dc/đc=được, a=anh, e=em, c=chị"],
    ["stk=số tài khoản, cod=thanh toán khi nhận hàng, bh=bảo hành"],
    [""],
    ["📌 LƯU Ý:"],
    ["- Hình ảnh phải là URL public (upload lên Imgur, Google Drive public)"],
    ["- Có thể tạo nhiều file Excel theo chủ đề (giá, sản phẩm, chính sách...)"],
    ["- Sau khi upload, nhấn 'Reload dữ liệu' trên Admin Panel"],
]

for row_idx, row_data in enumerate(guide_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_guide.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="4472C4")
        elif row_idx == 4:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        elif row_idx in [5, 6, 7, 8, 9]:
            cell.border = thin_border

ws_guide.column_dimensions['A'].width = 20
ws_guide.column_dimensions['B'].width = 60
ws_guide.column_dimensions['C'].width = 15

# ============ SHEET 2: Dữ liệu mẫu - Giá cả ============
ws_price = wb.create_sheet("Giá cả")

price_data = [
    ["câu hỏi", "câu trả lời", "hình ảnh", "từ khóa", "danh mục"],
    ["Giá sản phẩm bao nhiêu?", "Dạ giá sản phẩm dao động từ 100k - 500k tùy loại ạ. Anh/chị muốn xem sản phẩm nào để em báo giá chính xác ạ? 😊", "", "giá, tiền, bao nhiêu, bn", "Giá cả"],
    ["Giá bao nhiêu vậy shop?", "Dạ anh/chị cho em xin tên sản phẩm để em báo giá chính xác ạ 🌸", "", "giá, bao nhiêu, shop", "Giá cả"],
    ["Sp này bn tiền?", "Dạ sản phẩm này giá 150.000đ ạ. Mua từ 3 cái em giảm 10% luôn nha! 🎉", "", "sp, bn, tiền, giá", "Giá cả"],
    ["Có giảm giá không?", "Dạ có ạ! Hiện tại shop đang có chương trình:\n- Mua 2 giảm 5%\n- Mua 3 giảm 10%\n- Đơn từ 500k freeship ạ 🎁", "", "giảm giá, khuyến mãi, sale", "Giá cả"],
    ["Giá sỉ bao nhiêu?", "Dạ giá sỉ từ 10 cái trở lên em sẽ có giá tốt hơn ạ. Anh/chị inbox số lượng để em báo giá sỉ nhé! 📦", "", "sỉ, buôn, số lượng lớn", "Giá cả"],
]

for row_idx, row_data in enumerate(price_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_price.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_price.column_dimensions[col].width = 25 if col != 'B' else 70

# ============ SHEET 3: Vận chuyển ============
ws_ship = wb.create_sheet("Vận chuyển")

ship_data = [
    ["câu hỏi", "câu trả lời", "hình ảnh", "từ khóa", "danh mục"],
    ["Ship bao nhiêu?", "Dạ phí ship như sau ạ:\n- Nội thành HCM/HN: 20k\n- Tỉnh khác: 30k\n- Đơn từ 500k: FREESHIP 🚚", "", "ship, phí ship, vc", "Vận chuyển"],
    ["Có ship COD không?", "Dạ có ship COD toàn quốc ạ! Anh/chị nhận hàng rồi thanh toán luôn nha 📦", "", "cod, thanh toán khi nhận", "Vận chuyển"],
    ["Mấy ngày nhận được hàng?", "Dạ thời gian giao hàng:\n- Nội thành: 1-2 ngày\n- Tỉnh khác: 2-4 ngày\nEm sẽ gửi mã vận đơn ngay khi đóng gói xong ạ! 🚀", "", "mấy ngày, bao lâu, giao hàng", "Vận chuyển"],
    ["Giao hàng bằng gì?", "Dạ shop giao qua:\n- Giao hàng nhanh (GHN)\n- Giao hàng tiết kiệm (GHTK)\n- J&T Express\nAnh/chị chọn bên nào cũng được ạ! 📬", "", "giao hàng, đơn vị vc", "Vận chuyển"],
    ["Có freeship không?", "Dạ đơn từ 500k sẽ được FREESHIP toàn quốc ạ! Hoặc anh/chị mua combo 3 món cũng free ship luôn nha 🎉", "", "freeship, miễn phí ship, free vc", "Vận chuyển"],
]

for row_idx, row_data in enumerate(ship_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_ship.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_ship.column_dimensions[col].width = 25 if col != 'B' else 70

# ============ SHEET 4: Sản phẩm ============
ws_product = wb.create_sheet("Sản phẩm")

product_data = [
    ["câu hỏi", "câu trả lời", "hình ảnh", "từ khóa", "danh mục"],
    ["Có màu gì?", "Dạ sản phẩm có các màu:\n🔵 Xanh dương\n⚫ Đen\n⚪ Trắng\n🩷 Hồng\nAnh/chị thích màu nào ạ?", "", "màu, color, màu sắc", "Sản phẩm"],
    ["Có size gì?", "Dạ có đủ size từ S đến XXL ạ:\n- S: 40-50kg\n- M: 50-60kg\n- L: 60-70kg\n- XL: 70-80kg\nAnh/chị nặng bao nhiêu để em tư vấn size ạ? 📏", "", "size, sz, kích thước", "Sản phẩm"],
    ["Chất liệu gì?", "Dạ sản phẩm làm từ chất liệu cao cấp, mềm mại, thoáng mát ạ. Cam kết chất lượng như hình nha! ✨", "", "chất liệu, chất, vải", "Sản phẩm"],
    ["Hàng có sẵn không?", "Dạ có sẵn ạ! Anh/chị đặt hôm nay mai em gửi luôn nha 🚀", "", "có sẵn, còn hàng, hết hàng", "Sản phẩm"],
    ["Cho xem hình thật được không?", "Dạ đây là hình thật 100% ạ. Shop cam kết giao đúng như hình, không đúng hoàn tiền ạ! 📸", "", "hình thật, hình thực tế, ảnh thật", "Sản phẩm"],
]

for row_idx, row_data in enumerate(product_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_product.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_product.column_dimensions[col].width = 25 if col != 'B' else 70

# ============ SHEET 5: Chính sách ============
ws_policy = wb.create_sheet("Chính sách")

policy_data = [
    ["câu hỏi", "câu trả lời", "hình ảnh", "từ khóa", "danh mục"],
    ["Có đổi trả không?", "Dạ có ạ! Shop hỗ trợ đổi trả trong 7 ngày nếu:\n- Lỗi từ nhà sản xuất\n- Giao sai sản phẩm\n- Sản phẩm bị hư hỏng\nAnh/chị yên tâm mua sắm nha! 🛡️", "", "đổi trả, hoàn tiền, bảo đảm", "Chính sách"],
    ["Bảo hành bao lâu?", "Dạ sản phẩm được bảo hành 12 tháng ạ. Trong thời gian bảo hành nếu có lỗi em đổi mới miễn phí nha! 🔧", "", "bảo hành, bh, warranty", "Chính sách"],
    ["Thanh toán như thế nào?", "Dạ anh/chị có thể thanh toán:\n💳 Chuyển khoản trước\n💵 COD (nhận hàng trả tiền)\n\nSTK: 1234567890\nNgân hàng: Vietcombank\nChủ TK: Nguyễn Văn A", "", "thanh toán, ck, chuyển khoản, stk", "Chính sách"],
    ["Có hóa đơn không?", "Dạ có ạ! Shop xuất hóa đơn đầy đủ. Anh/chị cần hóa đơn VAT cứ nói em nhé 📝", "", "hóa đơn, bill, vat", "Chính sách"],
    ["Cam kết gì?", "Dạ shop cam kết:\n✅ Hàng chuẩn 100% như hình\n✅ Đổi trả 7 ngày\n✅ Bảo hành 12 tháng\n✅ Giao hàng đúng hẹn\nKhông đúng hoàn tiền ạ! 💯", "", "cam kết, đảm bảo, uy tín", "Chính sách"],
]

for row_idx, row_data in enumerate(policy_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_policy.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_policy.column_dimensions[col].width = 25 if col != 'B' else 70

# ============ SHEET 6: Chào hỏi ============
ws_greet = wb.create_sheet("Chào hỏi")

greet_data = [
    ["câu hỏi", "câu trả lời", "hình ảnh", "từ khóa", "danh mục"],
    ["Chào shop", "Dạ chào anh/chị ạ! 👋 Em là chatbot tư vấn của shop. Anh/chị cần hỗ trợ gì ạ? 😊", "", "chào, hello, hi, alo", "Chào hỏi"],
    ["Alo", "Dạ em nghe ạ! Anh/chị cần tư vấn sản phẩm nào ạ? 📱", "", "alo, a lô", "Chào hỏi"],
    ["Shop ơi", "Dạ em đây ạ! Anh/chị cần gì ạ? 🌸", "", "shop ơi, shop", "Chào hỏi"],
    ["Cảm ơn", "Dạ không có gì ạ! 🙏 Cảm ơn anh/chị đã tin tưởng shop. Chúc anh/chị ngày vui vẻ nha! 💕", "", "cảm ơn, thanks, tks", "Chào hỏi"],
    ["Tạm biệt", "Dạ tạm biệt anh/chị! 👋 Hẹn gặp lại, có gì cứ inbox shop nha! 🌟", "", "tạm biệt, bye, goodbye", "Chào hỏi"],
]

for row_idx, row_data in enumerate(greet_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_greet.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_greet.column_dimensions[col].width = 25 if col != 'B' else 70

# Save
output_path = "/home/claude/fb-chatbot/data/chatbot_template.xlsx"
wb.save(output_path)
print(f"✅ Đã tạo template: {output_path}")
